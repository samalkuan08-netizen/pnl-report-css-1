# -*- coding: utf-8 -*-
import io, tempfile, os
import streamlit as st
import openpyxl
import css_engine as eng
import recalc_native

st.set_page_config(page_title='P&L агент — CSS', layout='centered')
st.title('📊 P&L агент — CSS')
st.caption('1С ОСВ → единый лист P&L по блокам проектов. '
           'Можно загрузить несколько месяцев сразу. '
           'Новые статьи расходов >300 000 ₸ добавляются отдельной строкой автоматически.')

MONTHS=['Январь','Февраль','Март','Апрель','Май','Июнь',
        'Июль','Август','Сентябрь','Октябрь','Ноябрь','Декабрь']
COL={m:openpyxl.utils.get_column_letter(3+i) for i,m in enumerate(MONTHS)}

osv_files=st.file_uploader(
    'Выгрузки ОСВ из 1С — по одной на месяц (можно выбрать сразу несколько)',
    type=['xlsx'], accept_multiple_files=True)

start_month=st.selectbox('Первый месяц (файлы распределятся по месяцам по порядку)', MONTHS)

with st.expander('➕ Дозаполнить существующий отчёт (не потерять прежние месяцы)'):
    base_file=st.file_uploader('Текущий отчёт P&L', type=['xlsx'], key='base')

if osv_files:
    s=MONTHS.index(start_month)
    st.write('**Распределение по месяцам:**')
    for i,uf in enumerate(osv_files):
        m=MONTHS[min(s+i,11)]
        st.write(f'• `{uf.name}` → **{m}** (колонка {COL[m]})')

st.divider()

def _clean_month(ws, col):
    idx=openpyxl.utils.column_index_from_string(col)
    for r in range(3, ws.max_row+1):
        c=ws.cell(row=r,column=idx)
        if not (isinstance(c.value,str) and str(c.value).startswith('=')):
            c.value=None

if st.button('▶️ Сформировать отчёт', type='primary', disabled=not osv_files):
    try:
        if base_file is not None:
            tmpl_bytes=base_file.getvalue()
        else:
            with open(os.path.join(os.path.dirname(__file__),'template.xlsx'),'rb') as f:
                tmpl_bytes=f.read()
        work=tempfile.NamedTemporaryFile(delete=False,suffix='.xlsx').name
        with open(work,'wb') as f: f.write(tmpl_bytes)

        s=MONTHS.index(start_month)
        flags_all=[]
        for i,uf in enumerate(osv_files):
            m=MONTHS[min(s+i,11)]; col=COL[m]
            osv_tmp=tempfile.NamedTemporaryFile(delete=False,suffix='.xlsx').name
            with open(osv_tmp,'wb') as f: f.write(uf.getvalue())
            leaves=eng.parse_1c(osv_tmp)
            wb=openpyxl.load_workbook(work); ws=wb[eng.SHEET]; _clean_month(ws,col); wb.save(work)
            out=tempfile.NamedTemporaryFile(delete=False,suffix='.xlsx').name
            flags,log=eng.fill(work,leaves,out,month_col=col)
            work=out
            flags_all+=[(m,)+tuple(x) for x in flags]
            st.info(f'✅ {uf.name} → «{m}»: статей обработано {len(log)}')

        recalc_native.recalc_native(work)
        with open(work,'rb') as f: data=f.read()
        st.success('Готово!')
        st.download_button('⬇️ Скачать отчёт P&L', data=data,
                           file_name='CSS_P_L_отчёт.xlsx',
                           mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

        new_articles=[f for f in flags_all if len(f)>=5 and 'НОВАЯ' in str(f[4])]
        if new_articles:
            st.warning('🆕 Добавлены новые строки статей (>300 000 ₸):')
            for f in new_articles:
                st.write(f'• {f[0]}: «{f[2]}» — {f[3]:,.0f} ₸ (в месяцах без этой статьи ячейка пустая)')
        other=[f for f in flags_all if f not in new_articles]
        if other:
            with st.expander(f'⚠️ Требуют внимания ({len(other)})'):
                for f in other: st.write('•', f)
    except Exception as e:
        st.error(f'Ошибка: {e}')
        import traceback; st.code(traceback.format_exc())

with st.expander('ℹ️ Как работает разноска'):
    st.markdown('''
- **Доход** (6ххх) → блок «ДОХОД ОТ РЕАЛИЗАЦИИ» по проекту; вознаграждения (61хх) → «Доходы от финансовой деятельности»; курсовая (62хх) → «Прочие доходы».
- **Прямые затраты** (7010) → блок проекта. Налоги по ЗП = ОПВ+ОСМС+соц.отчисления+соц.налог. «Аренда судна» → строка судна.
- **Административные** (7210/7212) → блок «Административные расходы».
- **Прочие расходы** (74хх) → строка «Прочие расходы».
- **Новая статья**: **>300 000 ₸** — отдельной строкой в блок (в месяцах без неё ячейка пустая); **≤300 000 ₸** — в «Прочие» блока.
- Доход тянется из 1С; ручные источники добавляются бухгалтером отдельно.
''')
