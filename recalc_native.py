# -*- coding: utf-8 -*-
"""Лёгкий пересчёт формул отчёта без внешних зависимостей (только openpyxl).
Поддерживает: ссылки на ячейки (в т.ч. 'Лист'!C5), SUM(диапазон), IFERROR(x,y),
арифметику + - * / и числовые константы. Итеративно до стабилизации."""
import re, openpyxl
from openpyxl.utils import range_boundaries, get_column_letter

def recalc_native(path, max_iter=25):
    wb=openpyxl.load_workbook(path)
    sheets={ws.title:ws for ws in wb.worksheets}
    # значения: (sheet,cell)->число/строка; формулы: (sheet,cell)->строка формулы
    val={}; form={}
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if c.value is None: continue
                if isinstance(c.value,str) and c.value.startswith('='):
                    form[(ws.title,c.coordinate)]=c.value[1:]
                elif isinstance(c.value,(int,float)):
                    val[(ws.title,c.coordinate)]=float(c.value)
                else:
                    val[(ws.title,c.coordinate)]=c.value

    ref_re=re.compile(r"(?:'([^']+)'|([A-Za-z0-9 _\.&]+?))?!?\$?([A-Z]{1,3})\$?(\d+)")
    cell_re=re.compile(r"^\$?([A-Z]{1,3})\$?(\d+)$")

    def getval(sheet,coord):
        k=(sheet,coord)
        if k in val:
            v=val[k]; return v if isinstance(v,(int,float)) else 0.0
        if k in form: return None  # ещё не посчитано
        return 0.0

    def expand_sum(expr,cursheet):
        def repl(m):
            inside=m.group(1)
            tot=0.0
            for part in inside.split(','):
                part=part.strip()
                rng=re.match(r"(?:'([^']+)'!|([A-Za-z0-9 _\.&]+)!)?(\$?[A-Z]{1,3}\$?\d+):(\$?[A-Z]{1,3}\$?\d+)",part)
                if rng:
                    sh=rng.group(1) or rng.group(2) or cursheet
                    a=rng.group(3).replace('$',''); b=rng.group(4).replace('$','')
                    c1,r1,c2,r2=range_boundaries(f'{a}:{b}')
                    for rr in range(r1,r2+1):
                        for cc in range(c1,c2+1):
                            v=getval(sh,f'{get_column_letter(cc)}{rr}')
                            if v is None: return None
                            tot+=v
                else:
                    cm=re.match(r"(?:'([^']+)'!|([A-Za-z0-9 _\.&]+)!)?(\$?[A-Z]{1,3}\$?\d+)",part)
                    if cm:
                        sh=cm.group(1) or cm.group(2) or cursheet
                        v=getval(sh,cm.group(3).replace('$',''))
                        if v is None: return None
                        tot+=v
            return repr(tot)
        prev=None
        while prev!=expr:
            prev=expr
            expr=re.sub(r"SUM\(([^()]*)\)",lambda m:(lambda r:r if r is not None else 'None')(repl(m)),expr,flags=re.I)
            if 'None' in expr: return None
        return expr

    def sub_refs(expr,cursheet):
        # ссылки 'Лист'!C5 или Лист!C5 или C5
        def repl(m):
            sh=m.group(1) or m.group(2)
            col=m.group(3); rn=m.group(4)
            sheet=sh.strip() if sh else cursheet
            if sheet not in sheets: sheet=cursheet
            v=getval(sheet,f'{col}{rn}')
            return 'None' if v is None else repr(v)
        return ref_re.sub(repl,expr)

    def evaluate(expr,cursheet):
        # IFERROR(x,y)
        def iferr(m):
            inner=m.group(1)
            # делим аргументы по верхнеуровневой запятой
            depth=0; arg=''; args=[]
            for ch in inner:
                if ch=='(' : depth+=1
                if ch==')' : depth-=1
                if ch==',' and depth==0: args.append(arg); arg=''
                else: arg+=ch
            args.append(arg)
            x=evaluate(args[0],cursheet)
            if x is None: return 'None'
            try:
                r=eval_arith(x)
                return repr(r)
            except Exception:
                fb=args[1].strip() if len(args)>1 else '0'
                fb=fb.strip('"')
                return repr(fb) if fb else "''"
        prev=None
        while prev!=expr:
            prev=expr
            expr=re.sub(r"IFERROR\(([^()]*(?:\([^()]*\)[^()]*)*)\)",iferr,expr,flags=re.I)
        expr=expand_sum(expr,cursheet)
        if expr is None: return None
        expr=sub_refs(expr,cursheet)
        if 'None' in expr: return None
        return expr

    def eval_arith(s):
        if isinstance(s,(int,float)): return s
        s=str(s).strip()
        if s=='' or s=="''": return ''
        if not re.fullmatch(r"[-+*/().eE0-9 ]+",s):
            # строка-результат
            return s.strip("'\"")
        return eval(s,{"__builtins__":{}},{})

    # итеративно
    for _ in range(max_iter):
        changed=False
        for k,f in form.items():
            if k in val: continue
            r=evaluate(f,k[0])
            if r is None: continue
            try: rv=eval_arith(r)
            except Exception: continue
            val[k]=rv; changed=True
        if not changed: break

    # формулы оставляем в ячейках; сохраняем файл
    wb.calculation.fullCalcOnLoad=True
    wb.save(path)

    # вписываем посчитанные значения как КЭШ формул (formula + <v>), чтобы файл
    # показывал число сразу в любой программе, но при этом хранил живую формулу
    written=_inject_cached_values(path, {k:v for k,v in val.items() if k in form})
    return written


def _inject_cached_values(path, valmap):
    """Добавляет кэш <v>значение</v> в формульные ячейки прямо в XML xlsx."""
    import zipfile, shutil, os
    from xml.etree import ElementTree as ET
    MAIN='http://schemas.openxmlformats.org/spreadsheetml/2006/main'
    REL='http://schemas.openxmlformats.org/officeDocument/2006/relationships'
    PKGREL='http://schemas.openxmlformats.org/package/2006/relationships'
    ET.register_namespace('', MAIN)

    zin=zipfile.ZipFile(path,'r')
    names=zin.namelist()
    # карта: имя листа -> файл xl/worksheets/sheetN.xml
    wbroot=ET.fromstring(zin.read('xl/workbook.xml'))
    relroot=ET.fromstring(zin.read('xl/_rels/workbook.xml.rels'))
    rid2tgt={r.get('Id'):r.get('Target') for r in relroot}
    title2file={}
    for sh in wbroot.iter('{%s}sheet'%MAIN):
        rid=sh.get('{%s}id'%REL); tgt=rid2tgt.get(rid)
        if tgt:
            tgt=tgt.lstrip('/')
            if not tgt.startswith('xl/'): tgt='xl/'+tgt
            title2file[sh.get('name')]=tgt
    # значения по файлам-листам: {файл: {coord: value}}
    byfile={}
    for (title,coord),v in valmap.items():
        f=title2file.get(title)
        if f: byfile.setdefault(f,{})[coord]=v

    written=0
    newdata={}
    for fname in names:
        data=zin.read(fname)
        if fname in byfile:
            root=ET.fromstring(data)
            coords=byfile[fname]
            for c in root.iter('{%s}c'%MAIN):
                coord=c.get('r')
                if coord in coords:
                    fel=c.find('{%s}f'%MAIN)
                    if fel is not None:
                        vel=c.find('{%s}v'%MAIN)
                        if vel is None:
                            vel=ET.SubElement(c,'{%s}v'%MAIN)
                        val=coords[coord]
                        vel.text=(repr(float(val)) if isinstance(val,(int,float)) else str(val))
                        # порядок: <f> перед <v>
                        c.remove(vel); c.append(vel)
                        written+=1
            data=ET.tostring(root,xml_declaration=True,encoding='UTF-8')
        newdata[fname]=data
    zin.close()
    tmp=path+'.tmp'
    with zipfile.ZipFile(tmp,'w',zipfile.ZIP_DEFLATED) as zout:
        for fname in names:
            zout.writestr(fname,newdata[fname])
    shutil.move(tmp,path)
    return written
